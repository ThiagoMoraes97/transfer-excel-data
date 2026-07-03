import io
import re
import unittest
import zipfile

from openpyxl import Workbook, load_workbook

from web_app import _scan_sheet_columns


def _build_workbook_bytes():
    wb = Workbook()
    ws = wb.active
    ws.title = "Origem"

    ws["A1"] = "RELATORIO DE ATENDIMENTO MAIO"
    ws.append([])
    ws.append(
        [
            "Data de Competencia",
            "Data de Pagamento",
            "Valor",
            "Categoria",
            "Descricao",
            "Cliente",
            "Centro de Custo",
            "Observacoes",
        ]
    )
    ws.append(
        [
            "2026-05-04",
            "2026-05-04",
            50,
            "Cardiologia",
            "Dinheiro",
            "Robson",
            "",
            "Medic Beneficios",
        ]
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _with_broken_dimension(workbook_bytes: bytes) -> bytes:
    input_buffer = io.BytesIO(workbook_bytes)
    output_buffer = io.BytesIO()

    with zipfile.ZipFile(input_buffer, "r") as zin, zipfile.ZipFile(output_buffer, "w") as zout:
        for info in zin.infolist():
            content = zin.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                content = re.sub(
                    br'<dimension ref="[^"]+"',
                    b'<dimension ref="A1:A1"',
                    content,
                    count=1,
                )
            zout.writestr(info, content)

    return output_buffer.getvalue()


class ScanSheetColumnsTests(unittest.TestCase):
    def test_detects_headers_when_real_header_starts_on_third_row(self):
        workbook_bytes = _build_workbook_bytes()
        wb = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)

        columns = _scan_sheet_columns(wb["Origem"])

        self.assertEqual(8, len(columns))
        self.assertEqual("Data de Competencia", columns[0]["header"])
        self.assertEqual("2026-05-04", columns[0]["sample"])
        self.assertEqual("Observacoes", columns[7]["header"])

    def test_detects_columns_even_with_incorrect_dimension_metadata(self):
        workbook_bytes = _with_broken_dimension(_build_workbook_bytes())
        wb = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)

        columns = _scan_sheet_columns(wb["Origem"])

        self.assertEqual(8, len(columns))
        self.assertEqual("Categoria", columns[3]["header"])


if __name__ == "__main__":
    unittest.main()
