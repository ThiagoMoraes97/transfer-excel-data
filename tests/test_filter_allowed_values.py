import io
import unittest

from openpyxl import Workbook, load_workbook

from web_app import _copy_data, _matches_filter


def _workbook_bytes(sheet_name, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    for row in rows:
        ws.append(row)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class AllowedValuesFilterTests(unittest.TestCase):
    def test_not_in_list_excludes_values_outside_allowed_list(self):
        self.assertFalse(_matches_filter("Dinheiro", "not_in_list", "Dinheiro, Boleto"))
        self.assertFalse(_matches_filter(" boleto ", "not_in_list", "Dinheiro, Boleto"))
        self.assertTrue(_matches_filter("Pix", "not_in_list", "Dinheiro, Boleto"))
        self.assertTrue(_matches_filter("", "not_in_list", "Dinheiro, Boleto"))

    def test_can_keep_only_cash_and_boleto_while_excluding_zero_values(self):
        source_bytes = _workbook_bytes(
            "Origem",
            [
                ["Descricao", "Forma de Pagamento", "Valor"],
                ["Linha dinheiro", "Dinheiro", 100],
                ["Linha boleto", "Boleto", 50],
                ["Linha pix", "Pix", 80],
                ["Linha dinheiro zero", "Dinheiro", 0],
            ],
        )
        dest_bytes = _workbook_bytes("Destino", [["Descricao"]])

        output, copied_rows, copied_cells = _copy_data(
            source_bytes=source_bytes,
            source_filename="origem.xlsx",
            source_sheet="Origem",
            source_start_row=2,
            dest_bytes=dest_bytes,
            dest_filename="destino.xlsx",
            dest_sheet="Destino",
            dest_start_row=2,
            mappings=[{"sources": ["A"], "target": "A"}],
            separator=" - ",
            skip_empty_rows=True,
            filters=[
                {"column": "B", "operator": "not_in_list", "value": "Dinheiro, Boleto"},
                {"column": "C", "operator": "equals_zero", "value": ""},
            ],
        )

        wb_result = load_workbook(output)
        values = [wb_result["Destino"][f"A{row}"].value for row in range(2, 6)]

        self.assertEqual(["Linha dinheiro", "Linha boleto", None, None], values)
        self.assertEqual(2, copied_rows)
        self.assertEqual(2, copied_cells)


if __name__ == "__main__":
    unittest.main()
