import io
import unittest

from openpyxl import Workbook, load_workbook

from web_app import _copy_data


def _workbook_bytes(sheet_name, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    for row in rows:
        ws.append(row)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class CurrencyTransferTests(unittest.TestCase):
    def test_keeps_currency_text_when_option_is_disabled(self):
        source_bytes = _workbook_bytes("Origem", [["Valor"], ["R$50,00"]])
        dest_bytes = _workbook_bytes("Destino", [["Valor"]])

        output, copied_rows, copied_cells = _copy_data(
            source_bytes=source_bytes,
            source_filename="origem.xlsx",
            source_sheet="Origem",
            source_start_row=2,
            dest_bytes=dest_bytes,
            dest_filename="destino.xlsx",
            dest_sheet="Destino",
            dest_start_row=2,
            mappings=[{"sources": ["A"], "target": "A"}],
            separator=" - ",
            skip_empty_rows=True,
            remove_currency_symbol=False,
        )

        wb_result = load_workbook(output)
        self.assertEqual("R$50,00", wb_result["Destino"]["A2"].value)
        self.assertEqual(1, copied_rows)
        self.assertEqual(1, copied_cells)

    def test_removes_currency_symbol_when_option_is_enabled(self):
        source_bytes = _workbook_bytes("Origem", [["Valor"], ["R$50,00"]])
        dest_bytes = _workbook_bytes("Destino", [["Valor"]])

        output, _, _ = _copy_data(
            source_bytes=source_bytes,
            source_filename="origem.xlsx",
            source_sheet="Origem",
            source_start_row=2,
            dest_bytes=dest_bytes,
            dest_filename="destino.xlsx",
            dest_sheet="Destino",
            dest_start_row=2,
            mappings=[{"sources": ["A"], "target": "A"}],
            separator=" - ",
            skip_empty_rows=True,
            remove_currency_symbol=True,
        )

        wb_result = load_workbook(output)
        self.assertEqual(50.0, wb_result["Destino"]["A2"].value)


if __name__ == "__main__":
    unittest.main()
