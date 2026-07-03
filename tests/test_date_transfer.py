import io
import unittest
from datetime import datetime

from openpyxl import Workbook, load_workbook

from web_app import _copy_data, _matches_filter


def _workbook_bytes(sheet_name, configure_sheet):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    configure_sheet(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class DateTransferTests(unittest.TestCase):
    def test_month_filter_matches_excel_dates_and_pt_br_text_dates(self):
        self.assertFalse(_matches_filter(datetime(2026, 5, 4), "date_month_is", "5"))
        self.assertFalse(_matches_filter("05/05/2026", "date_month_is", "05"))
        self.assertTrue(_matches_filter(datetime(2026, 4, 30), "date_month_is", "5"))
        self.assertTrue(_matches_filter("10/06/2026", "date_month_is", "5"))
        self.assertTrue(_matches_filter("", "date_month_is", "5"))

    def test_can_keep_only_rows_from_selected_month(self):
        source_bytes = _workbook_bytes(
            "Origem",
            lambda ws: _fill_month_filter_source(ws),
        )
        dest_bytes = _workbook_bytes(
            "Destino",
            lambda ws: ws.append(["Descricao"]),
        )

        output, copied_rows, copied_cells = _copy_data(
            source_bytes=source_bytes,
            source_filename="origem.xlsx",
            source_sheet="Origem",
            source_start_row=2,
            dest_bytes=dest_bytes,
            dest_filename="destino.xlsx",
            dest_sheet="Destino",
            dest_start_row=2,
            mappings=[{"sources": ["A"], "target": "A"}],
            separator=" - ",
            skip_empty_rows=True,
            filters=[
                {"column": "B", "operator": "date_month_is", "value": "5"},
            ],
        )

        wb_result = load_workbook(output)
        values = [wb_result["Destino"][f"A{row}"].value for row in range(2, 6)]

        self.assertEqual(["Maio data Excel", "Maio texto", None, None], values)
        self.assertEqual(2, copied_rows)
        self.assertEqual(2, copied_cells)

    def test_preserves_source_date_display_format_for_single_column_mapping(self):
        source_bytes = _workbook_bytes(
            "Origem",
            lambda ws: _fill_source(ws),
        )
        dest_bytes = _workbook_bytes(
            "Destino",
            lambda ws: ws.append(["Data"]),
        )

        output, copied_rows, copied_cells = _copy_data(
            source_bytes=source_bytes,
            source_filename="origem.xlsx",
            source_sheet="Origem",
            source_start_row=2,
            dest_bytes=dest_bytes,
            dest_filename="destino.xlsx",
            dest_sheet="Destino",
            dest_start_row=2,
            mappings=[{"sources": ["A"], "target": "A"}],
            separator=" - ",
            skip_empty_rows=True,
        )

        wb_result = load_workbook(output)
        cell = wb_result["Destino"]["A2"]

        self.assertEqual(datetime(2026, 5, 4), cell.value)
        self.assertEqual("dd/mm/yyyy", cell.number_format)
        self.assertEqual(1, copied_rows)
        self.assertEqual(1, copied_cells)


def _fill_source(ws):
    ws.append(["Data"])
    ws["A2"] = datetime(2026, 5, 4)
    ws["A2"].number_format = "dd/mm/yyyy"


def _fill_month_filter_source(ws):
    ws.append(["Descricao", "Data"])
    ws.append(["Maio data Excel", datetime(2026, 5, 4)])
    ws.append(["Maio texto", "05/05/2026"])
    ws.append(["Abril", datetime(2026, 4, 30)])
    ws.append(["Junho texto", "10/06/2026"])


if __name__ == "__main__":
    unittest.main()
