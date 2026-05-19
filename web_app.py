import io
import json
import secrets
from datetime import datetime, timedelta

from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

app = Flask(__name__)

UPLOAD_CACHE = {}
CACHE_TTL_MINUTES = 30


def _is_empty(value):
    return value is None or str(value).strip() == ""


def _guess_keep_vba(filename: str) -> bool:
    return (filename or "").lower().endswith(".xlsm")


def _to_text(value):
    if value is None:
        return ""
    return str(value)


def _cleanup_cache():
    now = datetime.utcnow()
    expired = [
        token
        for token, payload in UPLOAD_CACHE.items()
        if now - payload["created_at"] > timedelta(minutes=CACHE_TTL_MINUTES)
    ]
    for token in expired:
        UPLOAD_CACHE.pop(token, None)


def _scan_sheet_columns(ws):
    used = set()
    samples = {}

    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    if max_row == 1 and max_col == 1 and _is_empty(ws.cell(row=1, column=1).value):
        return []

    headers = {}
    for col_idx in range(1, max_col + 1):
        headers[col_idx] = ws.cell(row=1, column=col_idx).value

    for row_index, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True),
        start=1,
    ):
        for col_idx, value in enumerate(row, start=1):
            if _is_empty(value):
                continue
            used.add(col_idx)
            if row_index > 1 and col_idx not in samples:
                samples[col_idx] = value

    columns = []
    for col_idx in sorted(used):
        letter = get_column_letter(col_idx)
        header = headers.get(col_idx)
        sample = samples.get(col_idx)
        columns.append(
            {
                "index": col_idx,
                "letter": letter,
                "header": _to_text(header),
                "sample": _to_text(sample),
                "label": f"{col_idx} ({letter})",
            }
        )

    return columns


def _workbook_info_from_bytes(content: bytes, filename: str):
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True, keep_vba=_guess_keep_vba(filename))
    sheets = wb.sheetnames
    if not sheets:
        raise ValueError("A planilha nao possui abas.")

    default_sheet = sheets[0]
    columns = _scan_sheet_columns(wb[default_sheet])

    return {
        "filename": filename,
        "sheets": sheets,
        "default_sheet": default_sheet,
        "columns": columns,
    }


def _get_cached_workbook(token: str):
    if token not in UPLOAD_CACHE:
        raise ValueError("Arquivo nao encontrado no cache. Envie novamente.")
    return UPLOAD_CACHE[token]


def _parse_mappings(mappings_raw):
    mappings = []
    for item in mappings_raw:
        sources = item.get("sources") or []
        sources = [str(x).strip().upper() for x in sources if str(x).strip()]
        if not sources:
            raise ValueError("Cada mapeamento precisa de pelo menos 1 coluna de origem.")
        if len(sources) > 2:
            raise ValueError("Cada mapeamento suporta no maximo 2 colunas de origem.")

        target = str(item.get("target") or "").strip().upper()
        if not target:
            raise ValueError("Cada mapeamento precisa de uma coluna de destino.")

        mappings.append({"sources": sources, "target": target})

    if not mappings:
        raise ValueError("Adicione ao menos 1 mapeamento.")
    return mappings


def _last_filled_row(ws, col_letter: str, start_row: int) -> int:
    for row in range(ws.max_row, start_row - 1, -1):
        value = ws[f"{col_letter}{row}"].value
        if not _is_empty(value):
            return row
    return start_row - 1


def _to_number(value):
    """Convert value to float, handling various formats like 0, 0.00, 0,00."""
    if value is None:
        return None
    try:
        # Handle string numbers with comma as decimal separator
        text = str(value).strip().replace(".", "").replace(",", ".")
        # If there are multiple dots, revert to original and try again
        if text.count(".") > 1:
            text = str(value).strip().replace(",", "")
        return float(text)
    except (ValueError, TypeError):
        return None


def _matches_filter(cell_value, operator, filter_value):
    """Check if cell value matches filter condition (should be excluded)."""
    text_value = str(cell_value).strip() if cell_value is not None else ""
    num_value = _to_number(cell_value)
    filter_num = _to_number(filter_value) if filter_value else None

    if operator == "empty":
        return text_value == ""
    if operator == "not_empty":
        return text_value != ""
    if operator == "equals":
        return text_value.lower() == filter_value.lower()
    if operator == "not_equals":
        return text_value.lower() != filter_value.lower()
    if operator == "contains":
        return filter_value.lower() in text_value.lower()
    if operator == "not_contains":
        return filter_value.lower() not in text_value.lower()
    if operator == "starts_with":
        return text_value.lower().startswith(filter_value.lower())
    if operator == "ends_with":
        return text_value.lower().endswith(filter_value.lower())

    # Numeric operators
    if operator == "equals_zero":
        if num_value is None:
            return False
        return abs(num_value) < 0.0001  # Treat very small numbers as zero
    if operator == "not_zero":
        if num_value is None:
            return False
        return abs(num_value) >= 0.0001
    if operator == "greater_than":
        if num_value is None or filter_num is None:
            return False
        return num_value > filter_num
    if operator == "less_than":
        if num_value is None or filter_num is None:
            return False
        return num_value < filter_num

    # List exclusion: exclude only if ALL words in cell are in the exclusion list
    if operator == "all_in_list":
        if not filter_value or not text_value:
            return False

        # Parse exclusion list from filter value
        exclusion_list = [p.strip().lower() for p in filter_value.split(",") if p.strip()]
        if not exclusion_list:
            return False

        # Parse cell values (split by comma)
        cell_parts = [p.strip().lower() for p in text_value.split(",") if p.strip()]
        if not cell_parts:
            return False

        # Check if ALL parts of the cell are in the exclusion list
        # If yes, exclude the row (return True)
        # If at least one part is NOT in the list, don't exclude (return False)
        return all(part in exclusion_list for part in cell_parts)

    return False


def _copy_data(
    source_bytes: bytes,
    source_filename: str,
    source_sheet: str,
    source_start_row: int,
    dest_bytes: bytes,
    dest_filename: str,
    dest_sheet: str,
    dest_start_row: int,
    mappings,
    separator: str,
    skip_empty_rows: bool,
    filters=None,
):
    wb_source = load_workbook(
        io.BytesIO(source_bytes),
        data_only=False,
        keep_vba=_guess_keep_vba(source_filename),
    )
    wb_dest = load_workbook(
        io.BytesIO(dest_bytes),
        keep_vba=_guess_keep_vba(dest_filename),
    )

    if source_sheet not in wb_source.sheetnames:
        raise ValueError(f"A aba de origem '{source_sheet}' nao existe.")
    if dest_sheet not in wb_dest.sheetnames:
        raise ValueError(f"A aba de destino '{dest_sheet}' nao existe.")

    ws_source = wb_source[source_sheet]
    ws_dest = wb_dest[dest_sheet]

    all_source_cols = sorted({col for m in mappings for col in m["sources"]})
    last_rows = [_last_filled_row(ws_source, col, source_start_row) for col in all_source_cols]
    last_row = max(last_rows) if last_rows else source_start_row - 1

    if last_row < source_start_row:
        buffer = io.BytesIO()
        wb_dest.save(buffer)
        buffer.seek(0)
        return buffer, 0, 0

    dest_row_cursor = dest_start_row
    copied_rows = 0
    copied_cells = 0

    for src_row in range(source_start_row, last_row + 1):
        row_values = {col: ws_source[f"{col}{src_row}"].value for col in all_source_cols}

        if skip_empty_rows and all(_is_empty(v) for v in row_values.values()):
            continue

        # Apply filters - skip row if any filter matches
        if filters:
            should_skip = False
            for f in filters:
                col = f["column"].upper()
                op = f["operator"]
                val = f.get("value", "")
                if col in row_values and _matches_filter(row_values[col], op, val):
                    should_skip = True
                    break
            if should_skip:
                continue

        if skip_empty_rows:
            dst_row = dest_row_cursor
            dest_row_cursor += 1
        else:
            dst_row = dest_start_row + (src_row - source_start_row)

        for mapping in mappings:
            src_values = [row_values[col] for col in mapping["sources"]]
            non_empty = [v for v in src_values if not _is_empty(v)]

            if not non_empty:
                final_value = None
            elif len(non_empty) == 1:
                final_value = non_empty[0]
            else:
                final_value = separator.join(str(v) for v in non_empty)

            ws_dest[f"{mapping['target']}{dst_row}"].value = final_value
            copied_cells += 1

        copied_rows += 1

    output = io.BytesIO()
    wb_dest.save(output)
    output.seek(0)
    return output, copied_rows, copied_cells


@app.get("/")
def index():
    return render_template("index.html")


@app.post("/api/upload")
def upload():
    try:
        _cleanup_cache()
        file = request.files.get("file")
        if file is None or file.filename == "":
            return jsonify({"ok": False, "error": "Selecione um arquivo Excel."}), 400

        content = file.read()
        if not content:
            return jsonify({"ok": False, "error": "Arquivo vazio."}), 400

        info = _workbook_info_from_bytes(content, file.filename)
        token = secrets.token_urlsafe(18)
        UPLOAD_CACHE[token] = {
            "filename": file.filename,
            "content": content,
            "created_at": datetime.utcnow(),
        }

        return jsonify({"ok": True, "token": token, "workbook": info})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@app.post("/api/sheet-columns")
def sheet_columns():
    try:
        data = request.get_json(force=True)
        token = (data.get("token") or "").strip()
        sheet_name = (data.get("sheet") or "").strip()

        if not token or not sheet_name:
            return jsonify({"ok": False, "error": "Token e nome da aba sao obrigatorios."}), 400

        cached = _get_cached_workbook(token)
        wb = load_workbook(
            io.BytesIO(cached["content"]),
            read_only=True,
            data_only=True,
            keep_vba=_guess_keep_vba(cached["filename"]),
        )

        if sheet_name not in wb.sheetnames:
            return jsonify({"ok": False, "error": f"A aba '{sheet_name}' nao existe."}), 400

        columns = _scan_sheet_columns(wb[sheet_name])
        return jsonify({"ok": True, "columns": columns})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@app.post("/api/transfer")
def transfer():
    try:
        data = request.get_json(force=True)

        source_token = (data.get("source_token") or "").strip()
        dest_token = (data.get("dest_token") or "").strip()
        source_sheet = (data.get("source_sheet") or "").strip()
        dest_sheet = (data.get("dest_sheet") or "").strip()

        source_start_row = int(data.get("source_start_row", 2))
        dest_start_row = int(data.get("dest_start_row", 2))
        separator = str(data.get("separator", " "))
        skip_empty_rows = bool(data.get("skip_empty_rows", True))

        if source_start_row < 1 or dest_start_row < 1:
            return jsonify({"ok": False, "error": "As linhas iniciais devem ser maiores que zero."}), 400

        mappings_raw = data.get("mappings") or []
        mappings = _parse_mappings(mappings_raw)

        source_cached = _get_cached_workbook(source_token)
        dest_cached = _get_cached_workbook(dest_token)

        filters_raw = data.get("filters") or []
        filters = []
        for f in filters_raw:
            filters.append({
                "column": str(f.get("column") or "").strip().upper(),
                "operator": str(f.get("operator") or "").strip(),
                "value": str(f.get("value") or "").strip(),
            })

        output, copied_rows, copied_cells = _copy_data(
            source_bytes=source_cached["content"],
            source_filename=source_cached["filename"],
            source_sheet=source_sheet,
            source_start_row=source_start_row,
            dest_bytes=dest_cached["content"],
            dest_filename=dest_cached["filename"],
            dest_sheet=dest_sheet,
            dest_start_row=dest_start_row,
            mappings=mappings,
            separator=separator,
            skip_empty_rows=skip_empty_rows,
            filters=filters if filters else None,
        )

        base_name = dest_cached["filename"]
        if "." in base_name:
            stem, ext = base_name.rsplit(".", 1)
            output_name = f"{stem}_atualizada.{ext}"
        else:
            output_name = f"{base_name}_atualizada.xlsx"

        response = send_file(
            output,
            as_attachment=True,
            download_name=output_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response.headers["X-Copied-Rows"] = str(copied_rows)
        response.headers["X-Copied-Cells"] = str(copied_cells)
        return response
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


if __name__ == "__main__":
    app.run(debug=True)
