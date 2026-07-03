import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, get_column_letter


def normalizar_coluna(valor: str) -> str:
    coluna = (valor or "").strip().upper()
    if not coluna:
        raise ValueError("Informe a coluna.")
    # Valida se a coluna existe no formato Excel (A, B, AA...)
    column_index_from_string(coluna)
    return coluna


def para_coluna_excel(valor: str) -> str:
    texto = (valor or "").strip().upper()
    if not texto:
        raise ValueError("Coluna vazia no mapeamento.")
    if texto.isdigit():
        indice = int(texto)
        if indice < 1:
            raise ValueError(f"Indice de coluna invalido: {texto}")
        return get_column_letter(indice)
    return normalizar_coluna(texto)


def parsear_mapeamentos(texto_mapeamentos: str):
    texto = (texto_mapeamentos or "").strip()
    if not texto:
        raise ValueError("Informe ao menos um mapeamento de colunas.")

    pares = [p.strip() for p in texto.replace(";", ",").split(",") if p.strip()]
    if not pares:
        raise ValueError("Mapeamento invalido.")

    mapeamentos = []
    for par in pares:
        if ":" not in par:
            raise ValueError(f"Mapeamento invalido: '{par}'. Use o formato origem:destino.")
        origem_raw, destino_raw = [x.strip() for x in par.split(":", 1)]
        colunas_origem = [x.strip() for x in origem_raw.split("+") if x.strip()]
        if not colunas_origem:
            raise ValueError(f"Mapeamento invalido: '{par}'.")
        if len(colunas_origem) > 2:
            raise ValueError(f"Mapeamento invalido: '{par}'. Use no maximo 2 colunas de origem por destino.")

        colunas_origem_excel = [para_coluna_excel(x) for x in colunas_origem]
        coluna_destino = para_coluna_excel(destino_raw)
        mapeamentos.append((colunas_origem_excel, coluna_destino))

    return mapeamentos


def ultima_linha_preenchida(ws, coluna: str, linha_inicial: int) -> int:
    for linha in range(ws.max_row, linha_inicial - 1, -1):
        valor = ws[f"{coluna}{linha}"].value
        if valor is not None and str(valor).strip() != "":
            return linha
    return linha_inicial - 1


def ultima_linha_preenchida_multiplas(ws, colunas_origem, linha_inicial: int) -> int:
    ultima = linha_inicial - 1
    for coluna in colunas_origem:
        ultima_coluna = ultima_linha_preenchida(ws, coluna, linha_inicial)
        if ultima_coluna > ultima:
            ultima = ultima_coluna
    return ultima


def valor_vazio(valor) -> bool:
    return valor is None or str(valor).strip() == ""


def combinar_valores(valores, separador: str):
    preenchidos = [v for v in valores if not valor_vazio(v)]
    if not preenchidos:
        return None
    if len(preenchidos) == 1:
        return preenchidos[0]
    return separador.join(str(v) for v in preenchidos)


def preservar_formato_data(origem_celula, destino_celula):
    if origem_celula is None or destino_celula is None:
        return
    if origem_celula.is_date and origem_celula.number_format:
        destino_celula.number_format = origem_celula.number_format


def copiar_colunas(
    arquivo_origem: str,
    arquivo_destino: str,
    planilha_origem: str,
    planilha_destino: str,
    mapeamentos,
    linha_origem_inicial: int,
    linha_destino_inicial: int,
    apenas_preenchidas: bool,
    separador: str,
    arquivo_saida: str,
):
    wb_origem = load_workbook(arquivo_origem, data_only=False)
    wb_destino = load_workbook(arquivo_destino)

    if planilha_origem not in wb_origem.sheetnames:
        raise ValueError(f"A planilha de origem '{planilha_origem}' nao existe.")
    if planilha_destino not in wb_destino.sheetnames:
        raise ValueError(f"A planilha de destino '{planilha_destino}' nao existe.")

    ws_origem = wb_origem[planilha_origem]
    ws_destino = wb_destino[planilha_destino]

    colunas_origem = sorted({coluna for origens, _ in mapeamentos for coluna in origens})
    ultima = ultima_linha_preenchida_multiplas(ws_origem, colunas_origem, linha_origem_inicial)
    if ultima < linha_origem_inicial:
        wb_destino.save(arquivo_saida)
        return 0, 0

    linhas_copiadas = 0
    celulas_copiadas = 0
    linha_destino_atual = linha_destino_inicial

    for linha_origem in range(linha_origem_inicial, ultima + 1):
        celulas_origem = {}
        valores = {}
        for coluna_origem in colunas_origem:
            celulas_origem[coluna_origem] = ws_origem[f"{coluna_origem}{linha_origem}"]
            valores[coluna_origem] = celulas_origem[coluna_origem].value

        if apenas_preenchidas:
            if all(valor_vazio(v) for v in valores.values()):
                continue
            destino_linha = linha_destino_atual
            linha_destino_atual += 1
        else:
            destino_linha = linha_destino_inicial + (linha_origem - linha_origem_inicial)

        for colunas_origem_map, coluna_destino in mapeamentos:
            # Apenas o valor e alterado. Formatos da planilha permanecem.
            valor_final = combinar_valores([valores[c] for c in colunas_origem_map], separador)
            celula_destino = ws_destino[f"{coluna_destino}{destino_linha}"]
            celula_destino.value = valor_final
            if len(colunas_origem_map) == 1:
                preservar_formato_data(celulas_origem[colunas_origem_map[0]], celula_destino)
            celulas_copiadas += 1
        linhas_copiadas += 1

    wb_destino.save(arquivo_saida)
    return linhas_copiadas, celulas_copiadas


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Transferir dados entre planilhas Excel")
        self.geometry("760x560")
        self.resizable(False, False)

        self.origem_var = tk.StringVar()
        self.destino_var = tk.StringVar()
        self.saida_var = tk.StringVar()

        self.planilha_origem_var = tk.StringVar(value="Sheet")
        self.planilha_destino_var = tk.StringVar(value="Sheet")

        self.mapeamentos_var = tk.StringVar(value="A:A")
        self.separador_var = tk.StringVar(value=" ")

        self.linha_origem_var = tk.StringVar(value="2")
        self.linha_destino_var = tk.StringVar(value="3")

        self.apenas_preenchidas_var = tk.BooleanVar(value=True)

        self._montar_layout()

    def _montar_layout(self):
        frame = ttk.Frame(self, padding=16)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="Planilha 1 (origem):").grid(row=0, column=0, sticky="w")
        ttk.Entry(frame, textvariable=self.origem_var, width=72).grid(row=1, column=0, sticky="w")
        ttk.Button(frame, text="Selecionar", command=self._selecionar_origem).grid(row=1, column=1, padx=(8, 0))

        ttk.Label(frame, text="Planilha 2 (destino):").grid(row=2, column=0, sticky="w", pady=(10, 0))
        ttk.Entry(frame, textvariable=self.destino_var, width=72).grid(row=3, column=0, sticky="w")
        ttk.Button(frame, text="Selecionar", command=self._selecionar_destino).grid(row=3, column=1, padx=(8, 0))

        ttk.Label(frame, text="Arquivo de saida (planilha 2 atualizada):").grid(row=4, column=0, sticky="w", pady=(10, 0))
        ttk.Entry(frame, textvariable=self.saida_var, width=72).grid(row=5, column=0, sticky="w")
        ttk.Button(frame, text="Salvar como", command=self._selecionar_saida).grid(row=5, column=1, padx=(8, 0))

        ttk.Separator(frame, orient="horizontal").grid(row=6, column=0, columnspan=2, sticky="ew", pady=16)

        ttk.Label(frame, text="Nome da aba de origem:").grid(row=7, column=0, sticky="w")
        ttk.Entry(frame, textvariable=self.planilha_origem_var, width=20).grid(row=8, column=0, sticky="w")

        ttk.Label(frame, text="Nome da aba de destino:").grid(row=7, column=1, sticky="w")
        ttk.Entry(frame, textvariable=self.planilha_destino_var, width=20).grid(row=8, column=1, sticky="w")

        ttk.Label(frame, text="Mapeamentos de coluna (origem:destino):").grid(row=9, column=0, sticky="w", pady=(12, 0))
        ttk.Entry(frame, textvariable=self.mapeamentos_var, width=72).grid(row=10, column=0, columnspan=2, sticky="w")

        ttk.Label(frame, text="Separador para combinacao (ex: -, /, |):").grid(row=11, column=0, sticky="w", pady=(12, 0))
        ttk.Entry(frame, textvariable=self.separador_var, width=20).grid(row=12, column=0, sticky="w")

        ttk.Label(frame, text="Linha inicial de origem (ex: 2):").grid(row=11, column=1, sticky="w", pady=(12, 0))
        ttk.Entry(frame, textvariable=self.linha_origem_var, width=10).grid(row=12, column=1, sticky="w")

        ttk.Label(frame, text="Linha inicial de destino (ex: 3):").grid(row=13, column=0, sticky="w", pady=(12, 0))
        ttk.Entry(frame, textvariable=self.linha_destino_var, width=10).grid(row=14, column=0, sticky="w")

        ttk.Checkbutton(
            frame,
            text="Copiar apenas linhas em que alguma coluna de origem esteja preenchida",
            variable=self.apenas_preenchidas_var,
        ).grid(row=15, column=0, columnspan=2, sticky="w", pady=(16, 0))

        ttk.Button(frame, text="Transferir dados", command=self._executar).grid(row=16, column=0, columnspan=2, pady=20)

        ajuda = (
            "Exemplos:\n"
            "Simples: A:A, D:B, F:C  |  Numeros: 2:1,4:2,6:3\n"
            "Combinado: B+C:A ou 2+4:1 (usa o separador definido no campo acima)"
        )
        ttk.Label(frame, text=ajuda).grid(row=17, column=0, columnspan=2, sticky="w")

    def _selecionar_origem(self):
        caminho = filedialog.askopenfilename(
            title="Selecione a planilha de origem",
            filetypes=[("Excel", "*.xlsx *.xlsm")],
        )
        if caminho:
            self.origem_var.set(caminho)
            if self.planilha_origem_var.get() == "Sheet":
                self.planilha_origem_var.set(self._aba_padrao(caminho))

    def _selecionar_destino(self):
        caminho = filedialog.askopenfilename(
            title="Selecione a planilha de destino",
            filetypes=[("Excel", "*.xlsx *.xlsm")],
        )
        if caminho:
            self.destino_var.set(caminho)
            if self.planilha_destino_var.get() == "Sheet":
                self.planilha_destino_var.set(self._aba_padrao(caminho))
            if not self.saida_var.get():
                base, ext = os.path.splitext(caminho)
                self.saida_var.set(f"{base}_atualizada{ext}")

    def _selecionar_saida(self):
        caminho = filedialog.asksaveasfilename(
            title="Salvar planilha de saida",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if caminho:
            self.saida_var.set(caminho)

    def _aba_padrao(self, caminho_arquivo: str) -> str:
        try:
            wb = load_workbook(caminho_arquivo, read_only=True)
            return wb.sheetnames[0]
        except Exception:
            return "Sheet"

    def _executar(self):
        try:
            arquivo_origem = self.origem_var.get().strip()
            arquivo_destino = self.destino_var.get().strip()
            arquivo_saida = self.saida_var.get().strip()

            if not arquivo_origem or not arquivo_destino or not arquivo_saida:
                raise ValueError("Selecione origem, destino e arquivo de saida.")

            planilha_origem = self.planilha_origem_var.get().strip()
            planilha_destino = self.planilha_destino_var.get().strip()
            if not planilha_origem or not planilha_destino:
                raise ValueError("Informe o nome das abas de origem e destino.")

            mapeamentos = parsear_mapeamentos(self.mapeamentos_var.get())
            separador = self.separador_var.get()

            linha_origem_inicial = int(self.linha_origem_var.get())
            linha_destino_inicial = int(self.linha_destino_var.get())

            if linha_origem_inicial < 1 or linha_destino_inicial < 1:
                raise ValueError("Linhas iniciais devem ser maiores que zero.")

            linhas_copiadas, celulas_copiadas = copiar_colunas(
                arquivo_origem=arquivo_origem,
                arquivo_destino=arquivo_destino,
                planilha_origem=planilha_origem,
                planilha_destino=planilha_destino,
                mapeamentos=mapeamentos,
                linha_origem_inicial=linha_origem_inicial,
                linha_destino_inicial=linha_destino_inicial,
                apenas_preenchidas=self.apenas_preenchidas_var.get(),
                separador=separador,
                arquivo_saida=arquivo_saida,
            )

            messagebox.showinfo(
                "Concluido",
                "Transferencia concluida com sucesso.\n"
                f"Linhas copiadas: {linhas_copiadas}\n"
                f"Celulas copiadas: {celulas_copiadas}\n\n"
                f"Arquivo gerado:\n{arquivo_saida}",
            )
        except Exception as erro:
            messagebox.showerror("Erro", str(erro))


if __name__ == "__main__":
    app = App()
    app.mainloop()
